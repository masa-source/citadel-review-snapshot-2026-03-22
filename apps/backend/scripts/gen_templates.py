"""
再帰的データ構造（report ルート）用のデモテンプレートと設計台ベースを生成する。
実行: apps/backend を cwd に python scripts/gen_templates.py

【Jinja2 準拠】load_report_context(FULL) が返すコンテキストは camelCase。
論理キー（ById / ByWorkerId / ByTagNumber / ByRole）と Ordered（リスト [1] 始まり）、Primary（1件目）を利用可能。
ハイフン付き UUID・数値キーはブラケット記法 ['key'] 必須。日時は datetime のため strftime 利用可。

推奨プレースホルダ:
  - 報告書・会社: reportTitle, company.name, controlNumber 等
  - 作成日: {{ createdAt.strftime('%Y/%m/%d') }}（本番は datetime。プレビューは JS 側で ISO 表示にフォールバック）
  - 作業者: reportWorkerPrimary.worker.name（1件目）
            reportWorkersByRole.leader.worker.name（役割キー）
            reportWorkersByWorkerId['uuid'].worker.name（安定・キー参照）※Jinja2 では [] 必須
            reportWorkersOrdered[1].worker.name（1番目）※リスト添字は数値 [1], [2]
  - 対象計器: targetInstrumentPrimary.tagNumber（1件目）
              targetInstrumentsByTagNumber['TAG-001'].tagNumber（デモは TAG-001 等）
              targetInstrumentsById['uuid'].tagNumber
              targetInstrumentsOrdered[1].tagNumber
  - 使用部品: usedPartPrimary.part.name, usedPartsOrdered[1].part.name, usedPartsById['id'].part.name
  - レポート customData（スキーマ駆動）: customData.year, customData.inspectionType 等（スキーマ定義に依存）
  - 対象計器 customData: targetInstrumentPrimary.customData.<キー>（スキーマ定義に依存）
  - 対象計器の表: targetInstrumentPrimary.tablesOrdered[1].roleKey（表1件目の役割キー）
                targetInstrumentPrimary.tablesOrdered[1].rows（Jinja で {% for row in ... %} と併用）
                targetInstrumentPrimary.tablesByRole.<roleKey>.rows（役割キーで参照）
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side


def _style_header(ws, row: int = 1) -> None:
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin = Side(style="thin")
    for c in range(1, ws.max_column + 1):
        ws.cell(row=row, column=c).border = Border(
            left=thin, right=thin, top=thin, bottom=thin
        )


def _template_dir() -> Path:
    return Path(__file__).resolve().parent.parent / "assets" / "template"


def write_report_demo_xlsx(out_path: Path) -> None:
    """再帰的オブジェクトグラフ用デモテンプレート（report ルート）。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "表紙"

    # タイトル行
    ws["A1"] = "作業報告書（デモ）"
    ws["A1"].font = Font(size=16, bold=True)
    ws.merge_cells("A1:D1")

    # プレースホルダ一覧（report ルート、camelCase・Jinja2 準拠）
    # キー参照: reportWorkersByWorkerId['uuid'], targetInstrumentsById['uuid'] 等（[] 必須）
    # Primary: 1件目。Ordered: reportWorkersOrdered[1] 等（リスト添字は数値）
    data = [
        ("項目", "プレースホルダ", "備考"),
        ("報告書タイトル", "{{ reportTitle }}", "レポートタイトル"),
        ("管理番号", "{{ controlNumber }}", ""),
        (
            "作成日",
            "{{ createdAt.strftime('%Y/%m/%d') }}",
            "本番は datetime。プレビューは ISO 表示",
        ),
        ("会社名", "{{ company.name }}", "発行会社"),
        (
            "作業者(1件目)",
            "{{ reportWorkerPrimary.worker.name }}",
            "役割: {{ reportWorkerPrimary.workerRole }}。安定: reportWorkersByRole.leader 等",
        ),
        (
            "作業者(役割キー)",
            "{{ reportWorkersByRole.leader.worker.name }}",
            "主担当=leader, 副担当=assistant",
        ),
        (
            "作業者(連番1番目)",
            "{{ reportWorkersOrdered[1].worker.name }}",
            "リスト添字は数値。2番目は [2]",
        ),
        (
            "対象計器(1件目)",
            "{{ targetInstrumentPrimary.instrument.name }}",
            "タグ: {{ targetInstrumentPrimary.tagNumber }}。安定: targetInstrumentsByTagNumber['TAG-001']",
        ),
        (
            "対象計器(タグ番号)",
            "{{ targetInstrumentsByTagNumber['TAG-001'].tagNumber }}",
            "同一レポート内でタグ番号が一意のとき。キーは [] 必須",
        ),
        (
            "取引先(役割キー)",
            "{{ reportClientsByRole.owner.company.name }}",
            "owner, contractor 等",
        ),
        (
            "メーカー(1件目)",
            "{{ targetInstrumentPrimary.instrument.company.name }}",
            "",
        ),
        (
            "使用部品(1件目)",
            "{{ usedPartPrimary.part.name }}",
            "数量: {{ usedPartPrimary.quantity }}。安定: usedPartsById['id'].part.name",
        ),
        ("部品メーカー(1件目)", "{{ usedPartPrimary.part.company.name }}", ""),
        (
            "レポート customData(年度)",
            "{{ customData.year }}",
            "スキーマ駆動。デモ: year, inspectionType",
        ),
        (
            "レポート customData(種別)",
            "{{ customData.inspectionType }}",
            "スキーマ定義の jsonSchema に依存",
        ),
        (
            "対象計器 customData",
            "{{ targetInstrumentPrimary.customData | default({}) }}",
            "スキーマ駆動。キーは targetInstrumentPrimary.customData.<key> で参照",
        ),
        (
            "対象計器の表(1件目・役割キー)",
            "{{ targetInstrumentPrimary.tablesOrdered[1].roleKey }}",
            "役割で参照: targetInstrumentPrimary.tablesByRole.測定結果.roleKey",
        ),
        (
            "対象計器の表の行(ループ)",
            "{{ targetInstrumentPrimary.tablesOrdered[1].rows }}",
            "Jinja: {% for row in targetInstrumentPrimary.tablesOrdered[1].rows %}",
        ),
    ]
    for i, row_data in enumerate(data, start=3):
        for j, val in enumerate(row_data, start=1):
            ws.cell(row=i, column=j, value=val)
    _style_header(ws, 3)
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 18

    # 2枚目: データ一覧イメージ（Primary = 1件目。安定させたい場合は targetInstrumentsById['uuid'] 等を使用）
    ws2 = wb.create_sheet("データ一覧")
    ws2["A1"] = "対象計器一覧"
    ws2["A1"].font = Font(bold=True, size=12)
    ws2["A2"] = "{{ targetInstrumentPrimary.instrument.name }}"
    ws2["B2"] = "{{ targetInstrumentPrimary.tagNumber }}"
    ws2["A3"] = "使用部品"
    ws2["A3"].font = Font(bold=True)
    ws2["A4"] = "{{ usedPartPrimary.part.name }}"
    ws2["B4"] = "{{ usedPartPrimary.part.company.name }}"
    ws2["C4"] = "{{ usedPartPrimary.quantity }}"
    ws2["A5"] = "レポート customData(年度)"
    ws2["A5"].font = Font(bold=True)
    ws2["A6"] = "{{ customData.year | default('') }}"
    ws2["B6"] = "{{ customData.inspectionType | default('') }}"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print("Generated:", out_path)


def write_base_xlsx(out_path: Path) -> None:
    """設計台で新規作成するためのベーステンプレート（最小限のプレースホルダ）。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "シート1"

    ws["A1"] = "タイトル"
    ws["A1"].font = Font(bold=True, size=12)
    ws["B1"] = "{{ reportTitle }}"
    ws["A2"] = "管理番号"
    ws["B2"] = "{{ controlNumber }}"
    ws["A3"] = "作成日"
    ws["B3"] = "{{ createdAt.strftime('%Y/%m/%d') }}"
    ws["A4"] = "発行会社"
    ws["B4"] = "{{ company.name }}"
    ws["A5"] = "レポート customData(例)"
    ws["B5"] = "{{ customData.year | default('') }}"
    # 1件目: reportWorkerPrimary。キー参照は reportWorkersByWorkerId['uuid'].worker.name（[] 必須）
    ws["A6"] = "作業者（1件目）"
    ws["B6"] = "{{ reportWorkerPrimary.worker.name }}"
    ws["A7"] = "（設計台でセルを追加し、プレースホルダを配置してください）"
    ws["A7"].font = Font(italic=True)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 48

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print("Generated:", out_path)


def main() -> None:
    base = _template_dir()
    write_report_demo_xlsx(base / "Report_Demo_1.xlsx")
    write_base_xlsx(base / "Base.xlsx")


if __name__ == "__main__":
    main()
