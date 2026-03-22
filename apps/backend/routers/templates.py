"""テンプレート管理 API（一覧・scan・sync-file・CRUD・設計台・再検証）。"""

import errno
import json
import logging
import shutil
import tempfile
from datetime import UTC, datetime
from io import BytesIO
from pathlib import Path
from uuid import UUID

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from openpyxl import load_workbook
from pydantic import ValidationError
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from database import get_session
from error_codes import (
    BACKUP_FAILED,
    FILE_IN_USE,
    FILE_MODIFIED_EXTERNALLY,
    FILE_NOT_FOUND,
)
from models import ReportFormatTemplate, ReportTemplate
from schemas import (
    GridUpdateBody,
    TemplateAutoGenerateReportItem,
    TemplateAutoGenerateResponse,
    TemplateAutoGenerateTemplateItem,
    TemplateRevalidateBody,
    TemplateScanMissingItem,
    TemplateScanResult,
    TemplateSyncFileInput,
    TemplateUpdate,
)
from services.ai_extractor import extract_data_from_text
from services.ai_report_builder import build_report_from_ai_data
from services.db_loader import LoadScope, load_report_context
from services.template_auto_placer import generate_auto_placeholders
from services.template_editor import load_grid, save_grid
from services.template_safety import verify_template_safety
from utils.excel_to_text import grid_to_text
from utils.paths import (
    get_assets_base,
    get_assets_templates_dir,
    get_valid_template_paths,
)
from utils.quarantine import quarantine_xlsx

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api", tags=["templates"])


def _is_file_in_use_error(e: BaseException) -> bool:
    if isinstance(e, PermissionError):
        return True
    msg = str(e).lower()
    return bool(
        getattr(e, "errno", None) in (errno.EACCES, errno.EBUSY, 32, 13)
        or "used by another process" in msg
        or "permission denied" in msg
        or "access is denied" in msg
        or "being used" in msg
    )


def _raise_file_in_use(e: BaseException) -> None:
    raise HTTPException(
        status_code=409,
        detail={
            "code": FILE_IN_USE,
            "message": "このファイルは別のアプリ（Excel など）で開かれている可能性があります。ファイルを閉じてから再試行してください。",
        },
    ) from e


@router.get("/templates")
async def get_templates(
    session: AsyncSession = Depends(get_session),
):
    """ReportTemplate（テンプレート部品）の一覧を返す。fileExists は実ファイルの存在を示す。"""
    base = get_assets_base()
    result = await session.execute(
        select(ReportTemplate).order_by(ReportTemplate.name, ReportTemplate.id)
    )
    templates = result.scalars().all()
    return [
        {
            "id": t.id,
            "name": t.name,
            "filePath": t.file_path,
            "fileExists": (base / t.file_path).exists() if t.file_path else False,
        }
        for t in templates
    ]


@router.get("/templates/scan", response_model=TemplateScanResult)
async def scan_templates(
    session: AsyncSession = Depends(get_session),
):
    """
    テンプレートディレクトリ（TEMPLATE_DIR）内の .xlsx と DB の file_path を比較し、
    DB にないファイル（新規）とディスクにない DB レコード（行方不明）を返す。
    """
    base = get_assets_base()
    templates_dir = get_assets_templates_dir()
    disk_rel_paths = set(get_valid_template_paths(base, templates_dir.name))

    result = await session.execute(
        select(ReportTemplate).order_by(ReportTemplate.name, ReportTemplate.id)
    )
    templates = result.scalars().all()
    db_paths = {t.file_path for t in templates if t.file_path}

    new_files = sorted(disk_rel_paths - db_paths)
    missing_from_disk = [
        TemplateScanMissingItem(id=t.id, file_path=t.file_path)
        for t in templates
        if t.file_path and not (base / t.file_path).exists()
    ]
    inconsistent = bool(new_files or missing_from_disk)

    return TemplateScanResult(
        inconsistent=inconsistent,
        new_files=new_files,
        missing_from_disk=missing_from_disk,
    )


@router.post("/templates/sync-file")
async def sync_file_template(
    body: TemplateSyncFileInput,
    session: AsyncSession = Depends(get_session),
):
    """
    指定した file_path（assets 基準の相対パス）のファイルを検疫し、
    通過した場合のみ DB に ReportTemplate を 1 件作成する。
    """
    file_path = body.file_path.strip()
    if not file_path:
        raise HTTPException(status_code=400, detail="file_path を指定してください。")

    base = get_assets_base()
    full_path = (base / file_path).resolve()
    if not str(full_path).startswith(str(base.resolve())):
        raise HTTPException(status_code=400, detail="無効な file_path です。")

    if not full_path.is_file():
        raise HTTPException(
            status_code=404,
            detail=f"ファイルが見つかりません: {file_path}",
        )

    result = quarantine_xlsx(file_path=full_path)
    if not result.ok:
        raise HTTPException(
            status_code=400,
            detail=result.message,
        )

    existing = await session.execute(
        select(ReportTemplate).where(ReportTemplate.file_path == file_path)
    )
    if existing.scalar_one_or_none():
        raise HTTPException(
            status_code=409,
            detail=f"このファイルは既に登録されています: {file_path}",
        )

    default_name = full_path.stem

    template = ReportTemplate(
        name=default_name,
        file_path=file_path,
        last_verified_mtime=full_path.stat().st_mtime,
    )
    session.add(template)
    await session.commit()
    await session.refresh(template)

    return {
        "id": template.id,
        "name": template.name,
        "filePath": template.file_path,
    }


@router.post("/templates")
async def create_template(
    file: UploadFile = File(...),
    name: str = Form(...),
    session: AsyncSession = Depends(get_session),
):
    """
    ブックバインダー：.xlsx を検疫し、通過したもののみ assets/templates に登録する（テンプレート部品として）。
    """
    from io import BytesIO

    if not file.filename:
        raise HTTPException(
            status_code=400,
            detail="ファイル名がありません。このファイルはお受けできません。",
        )

    content = await file.read()
    buf = BytesIO(content)
    result = quarantine_xlsx(
        file_obj=buf,
        filename=file.filename,
        file_size=len(content),
    )
    if not result.ok:
        raise HTTPException(
            status_code=400,
            detail=result.message,
        )

    templates_dir = get_assets_templates_dir()
    templates_dir.mkdir(parents=True, exist_ok=True)

    safe_filename = file.filename.replace(" ", "_")
    dest_path = templates_dir / safe_filename
    if dest_path.exists():
        raise HTTPException(
            status_code=400,
            detail=f"同名のファイルが既に存在します: {safe_filename}",
        )

    try:
        buf.seek(0)
        with open(dest_path, "wb") as f:
            shutil.copyfileobj(buf, f)
    except Exception as e:
        logger.exception("テンプレートファイル保存エラー: %s", e)
        raise HTTPException(
            status_code=500, detail="ファイルの保存に失敗しました。"
        ) from e

    template = ReportTemplate(
        name=name or dest_path.stem,
        file_path=f"{templates_dir.name}/{safe_filename}",
        last_verified_mtime=dest_path.stat().st_mtime,
    )
    session.add(template)
    try:
        await session.commit()
        await session.refresh(template)
    except Exception as e:
        await session.rollback()
        try:
            if dest_path.exists():
                dest_path.unlink()
        except OSError:
            pass
        logger.exception("テンプレートDB保存エラー: %s", e)
        raise HTTPException(
            status_code=500, detail="テンプレートの保存に失敗しました。"
        ) from e

    return {
        "id": template.id,
        "name": template.name,
        "filePath": template.file_path,
    }


@router.post(
    "/templates/auto-generate",
    response_model=TemplateAutoGenerateResponse,
)
async def auto_generate_template(
    file: UploadFile = File(...),
    name: str = Form(...),
    session: AsyncSession = Depends(get_session),
) -> TemplateAutoGenerateResponse:
    """
    テンプレート .xlsx をアップロードし、AI でレポート情報を抽出したうえで
    プレースホルダを自動配置し、テンプレート部品として登録する。
    """
    if not file.filename:
        raise HTTPException(
            status_code=400,
            detail="ファイル名がありません。このファイルはお受けできません。",
        )

    content = await file.read()
    buf = BytesIO(content)
    result = quarantine_xlsx(
        file_obj=buf,
        filename=file.filename,
        file_size=len(content),
    )
    if not result.ok:
        raise HTTPException(status_code=400, detail=result.message)

    safe_filename = (file.filename or "template.xlsx").replace(" ", "_")
    if not safe_filename.lower().endswith(".xlsx"):
        safe_filename = safe_filename + ".xlsx"

    temp_dir = Path(tempfile.mkdtemp(prefix="auto_gen_"))
    temp_path = temp_dir / safe_filename
    try:
        temp_path.write_bytes(content)

        grid_data = load_grid(temp_path)
        text = grid_to_text(grid_data)
        ai_data = await extract_data_from_text(text)
        report = await build_report_from_ai_data(session, ai_data)
        await session.flush()

        report_ctx = await load_report_context(session, report.id, scope=LoadScope.FULL)
        context_data = (
            report_ctx.model_dump(by_alias=True, mode="python") if report_ctx else {}
        )
        changes = generate_auto_placeholders(grid_data, context_data)
        save_grid(temp_path, changes, use_excel_instance=False)

        templates_dir = get_assets_templates_dir()
        templates_dir.mkdir(parents=True, exist_ok=True)
        dest_path = templates_dir / safe_filename
        if dest_path.exists():
            raise HTTPException(
                status_code=400,
                detail=f"同名のファイルが既に存在します: {safe_filename}",
            )
        shutil.copy2(temp_path, dest_path)
        rel_path = f"{templates_dir.name}/{safe_filename}"

        template = ReportTemplate(
            name=name or dest_path.stem,
            file_path=rel_path,
            last_verified_mtime=dest_path.stat().st_mtime,
        )
        session.add(template)
        await session.commit()
        await session.refresh(template)

        return TemplateAutoGenerateResponse(
            template=TemplateAutoGenerateTemplateItem(
                id=template.id,
                name=template.name,
                file_path=template.file_path,
            ),
            report=TemplateAutoGenerateReportItem(
                id=report.id,
                report_title=report.report_title or "",
            ),
        )
    except HTTPException:
        await session.rollback()
        raise
    except (ValueError, ValidationError) as e:
        await session.rollback()
        raise HTTPException(status_code=400, detail=str(e)) from e
    except (PermissionError, OSError) as e:
        await session.rollback()
        if _is_file_in_use_error(e):
            _raise_file_in_use(e)
        raise HTTPException(
            status_code=500,
            detail=f"自動生成に失敗しました。{e!s}",
        ) from e
    except Exception as e:
        await session.rollback()
        logger.exception("テンプレート自動生成エラー: %s", e)
        raise HTTPException(
            status_code=500,
            detail=f"テンプレートの自動生成に失敗しました。{e!s}",
        ) from e
    finally:
        try:
            if temp_path.exists():
                temp_path.unlink()
            if temp_dir.exists():
                temp_dir.rmdir()
        except OSError:
            pass


@router.delete("/templates/{template_id}")
async def delete_template(
    template_id: UUID,
    session: AsyncSession = Depends(get_session),
):
    """指定IDのテンプレート部品をDBとファイルシステムから削除する。紐づく ReportFormatTemplate も削除。"""
    result = await session.execute(
        select(ReportTemplate).where(ReportTemplate.id == template_id)
    )
    template = result.scalar_one_or_none()
    if not template:
        raise HTTPException(status_code=404, detail="テンプレートが見つかりません。")

    # 中継テーブルの参照を先に削除
    links_result = await session.execute(
        select(ReportFormatTemplate).where(
            ReportFormatTemplate.report_template_id == template_id
        )
    )
    for link in links_result.scalars().all():
        await session.delete(link)

    if template.file_path:
        file_path = get_assets_base() / template.file_path
        if file_path.exists():
            try:
                file_path.unlink()
            except Exception as e:
                if _is_file_in_use_error(e):
                    await session.rollback()
                    _raise_file_in_use(e)
                await session.rollback()
                logger.warning("テンプレートファイル削除エラー: %s", e)
                raise HTTPException(
                    status_code=500,
                    detail="テンプレートファイルの削除に失敗しました。",
                ) from e

    await session.delete(template)
    await session.commit()

    return {"ok": True, "message": f"テンプレート ID={template_id} を削除しました。"}


@router.put("/templates/{template_id}")
async def update_template(
    template_id: UUID,
    update_data: TemplateUpdate,
    session: AsyncSession = Depends(get_session),
):
    """指定IDのテンプレート部品のメタデータを更新する（ファイルは変更しない）。"""
    result = await session.execute(
        select(ReportTemplate).where(ReportTemplate.id == template_id)
    )
    template = result.scalar_one_or_none()
    if not template:
        raise HTTPException(status_code=404, detail="テンプレートが見つかりません。")

    if update_data.name is not None:
        template.name = update_data.name
    if update_data.file_path is not None:
        base = get_assets_base()
        full_path = (base / update_data.file_path).resolve()
        if not full_path.is_relative_to(base.resolve()):
            raise HTTPException(status_code=400, detail="無効な file_path です。")
        template.file_path = update_data.file_path

    await session.commit()
    await session.refresh(template)

    return {
        "id": template.id,
        "name": template.name,
        "filePath": template.file_path,
    }


@router.get("/templates/{template_id}/grid")
async def get_template_grid(
    template_id: UUID,
    session: AsyncSession = Depends(get_session),
):
    """
    簡易設計台：テンプレートの全シートのセル値（0-based グリッド）を返す。
    """
    result = await session.execute(
        select(ReportTemplate).where(ReportTemplate.id == template_id)
    )
    template = result.scalar_one_or_none()
    if not template:
        raise HTTPException(status_code=404, detail="テンプレートが見つかりません。")
    if not template.file_path:
        raise HTTPException(
            status_code=400, detail="テンプレートにファイルが紐づいていません。"
        )
    file_path = get_assets_base() / template.file_path
    if not file_path.exists():
        raise HTTPException(
            status_code=404, detail="テンプレートファイルが見つかりません。"
        )
    try:
        await verify_template_safety(file_path, template, session)
    except (PermissionError, OSError) as e:
        if _is_file_in_use_error(e):
            _raise_file_in_use(e)
        raise
    try:
        data = load_grid(file_path)
    except Exception as e:
        logger.exception("テンプレートグリッド読み込みエラー: %s", e)
        raise HTTPException(
            status_code=500, detail="テンプレートの読み込みに失敗しました。"
        ) from e

    current_sheet_names = [s["name"] for s in data.get("sheets", [])]
    stored_sheet_names: list[str] | None = None
    if template.sheet_names:
        try:
            stored_sheet_names = json.loads(template.sheet_names)
        except (json.JSONDecodeError, TypeError):
            stored_sheet_names = None
    if stored_sheet_names is None:
        # GET では永続化しない（副作用を避ける）。保存は revalidate 等の POST で行う。
        data["storedSheetNamesMissing"] = True
    elif set(stored_sheet_names) != set(current_sheet_names):
        data["sheetNameMismatch"] = True
        data["storedSheetNames"] = stored_sheet_names
        data["currentSheetNames"] = current_sheet_names
    return data


@router.post("/templates/{template_id}/revalidate")
async def revalidate_template(
    template_id: UUID,
    body: TemplateRevalidateBody,
    session: AsyncSession = Depends(get_session),
):
    """
    外部編集モード用。物理ファイルの存在確認・検疫を行い、問題なければ更新日時を返す。
    ファイルがロックされている場合（Excel で開いている等）は 400 を返す。
    ファイルが見つからない場合は new_file_path で再試行できる。
    """
    result = await session.execute(
        select(ReportTemplate).where(ReportTemplate.id == template_id)
    )
    template = result.scalar_one_or_none()
    if not template:
        raise HTTPException(status_code=404, detail="テンプレートが見つかりません。")

    base = get_assets_base()
    rel_path = (body.new_file_path or template.file_path or "").strip()
    if not rel_path:
        raise HTTPException(
            status_code=400, detail="テンプレートにファイルパスが紐づいていません。"
        )
    full_path = (base / rel_path).resolve()
    if not str(full_path).startswith(str(base.resolve())):
        raise HTTPException(status_code=400, detail="無効なファイルパスです。")

    if not full_path.is_file():
        raise HTTPException(
            status_code=404,
            detail={
                "code": FILE_NOT_FOUND,
                "message": "テンプレートファイルが見つかりません。",
                "currentFilePath": template.file_path or "",
            },
        )

    if not body.force_continue:
        backup_dir = base / "backup"
        try:
            backup_dir.mkdir(parents=True, exist_ok=True)
            backup_name = f"{template_id}_{full_path.name}.bak"
            backup_path = backup_dir / backup_name
            shutil.copy2(full_path, backup_path)
        except PermissionError as e:
            logger.warning("バックアップ作成に失敗（権限）: %s", full_path)
            raise HTTPException(
                status_code=409,
                detail={
                    "code": BACKUP_FAILED,
                    "message": "バックアップの作成に失敗しました。このまま続行しますか？",
                    "reason": "ファイルが Excel など別のプログラムで開かれている可能性があります。",
                },
            ) from e
        except OSError as e:
            reason = "権限やディスクの状態を確認してください。"
            if getattr(e, "errno", None) == errno.ENOSPC:
                reason = "ディスクの空き容量が不足している可能性があります。"
            elif getattr(e, "errno", None) in (errno.EACCES, 13):
                reason = "ファイルが Excel など別のプログラムで開かれている可能性があります。"
            logger.warning("再検証時のバックアップに失敗: %s", e)
            raise HTTPException(
                status_code=409,
                detail={
                    "code": BACKUP_FAILED,
                    "message": "バックアップの作成に失敗しました。このまま続行しますか？",
                    "reason": reason,
                },
            ) from e

    try:
        result = quarantine_xlsx(file_path=full_path)
    except (PermissionError, OSError) as e:
        if _is_file_in_use_error(e):
            logger.warning("テンプレートファイルがロックされています: %s", full_path)
            _raise_file_in_use(e)
        raise HTTPException(
            status_code=500,
            detail=f"ファイルの読み込みに失敗しました。{e!s}",
        ) from e

    if not result.ok:
        raise HTTPException(status_code=400, detail=result.message)

    mtime = full_path.stat().st_mtime
    template.last_verified_mtime = mtime
    try:
        wb = load_workbook(full_path, read_only=True)
        template.sheet_names = json.dumps(wb.sheetnames, ensure_ascii=False)
        wb.close()
    except Exception as e:
        logger.warning("再検証後のシート名取得に失敗（無視）: %s", e)
    if body.new_file_path and body.new_file_path.strip():
        template.file_path = body.new_file_path.strip()
    await session.commit()
    await session.refresh(template)

    last_modified = datetime.fromtimestamp(mtime, tz=UTC).isoformat()

    return {
        "ok": True,
        "filePath": template.file_path,
        "lastModified": last_modified,
    }


@router.post("/templates/{template_id}/grid")
async def update_template_grid(
    template_id: UUID,
    body: GridUpdateBody,
    session: AsyncSession = Depends(get_session),
):
    """
    簡易設計台：差分データを受け取り、該当セルの値のみ更新してファイルを上書き保存する。
    """
    try:
        result = await session.execute(
            select(ReportTemplate).where(ReportTemplate.id == template_id)
        )
        template = result.scalar_one_or_none()
        if not template:
            raise HTTPException(
                status_code=404, detail="テンプレートが見つかりません。"
            )
        if not template.file_path:
            raise HTTPException(
                status_code=400, detail="テンプレートにファイルが紐づいていません。"
            )
        file_path = get_assets_base() / template.file_path
        if not file_path.exists():
            raise HTTPException(
                status_code=404, detail="テンプレートファイルが見つかりません。"
            )
        if not body.force_overwrite and template.last_verified_mtime is not None:
            current_mtime = file_path.stat().st_mtime
            if current_mtime != template.last_verified_mtime:
                raise HTTPException(
                    status_code=409,
                    detail={
                        "code": FILE_MODIFIED_EXTERNALLY,
                        "message": "ファイルが外部で変更されています。",
                    },
                )
        changes = [
            {"sheetName": c.sheet_name, "row": c.row, "col": c.col, "value": c.value}
            for c in body.changes
        ]
        save_grid(file_path, changes, use_excel_instance=body.use_excel_instance)
        template.last_verified_mtime = file_path.stat().st_mtime
        session.add(template)
        await session.commit()
        return {"ok": True, "message": "保存しました。"}
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e)) from e
    except (PermissionError, OSError) as e:
        msg = (e.args[0] if e.args else str(e)).lower()
        if (
            getattr(e, "errno", None) in (errno.EACCES, errno.EBUSY, 32)
            or "used by another process" in msg
            or "permission denied" in msg
            or "access is denied" in msg
            or "being used" in msg
        ):
            raise HTTPException(
                status_code=409,
                detail={
                    "code": FILE_IN_USE,
                    "message": "このファイルは別のアプリ（Excel など）で開かれている可能性があります。ファイルを閉じてから再度保存してください。",
                },
            ) from e
        raise HTTPException(
            status_code=500,
            detail=f"テンプレートの保存に失敗しました。{e!s}",
        ) from e
    except Exception as e:
        logger.exception("テンプレートグリッド保存エラー: %s", e)
        raise HTTPException(
            status_code=500,
            detail=f"テンプレートの保存に失敗しました。{e!s}",
        ) from e
