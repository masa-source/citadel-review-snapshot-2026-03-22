"""
簡易設計台（Lite Drafting Table）用：openpyxl でテンプレートの「値」のみ読み書き。

- 読み込み: 各セルの値のみ抽出し、フロント用に 0-based のグリッドで返す。
  行・列の非表示情報（row_metadata, col_metadata）も返す。
- 保存: 指定されたセルの値のみ更新し、罫線・背景色・フォント等は維持する。
  数式セルへの書き込み時は数式をクリアして文字列として保存。結合セル「左上以外」への書き込みは拒否。
"""

from __future__ import annotations

import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def load_grid(file_path: Path) -> dict[str, Any]:
    """
    テンプレート .xlsx を開き、全シートのセル値を 0-based で返す。
    data_only=False で開くため、数式は数式のまま・プレースホルダはそのまま渡る。

    Returns:
        {
          "sheets": [
            {
              "name": "Sheet1",
              "data": [[cell_value, ...], ...],
              "mergeCells": [...],
              "row_metadata": [ { "hidden": bool }, ... ],
              "col_metadata": [ { "hidden": bool, "width": int }, ... ]
            },
            ...
          ]
        }
    """
    file_path = Path(file_path).resolve()
    if not file_path.exists():
        raise FileNotFoundError(f"テンプレートが見つかりません: {file_path}")

    wb = load_workbook(file_path, read_only=False, data_only=False)
    sheets_out = []

    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # openpyxl は 1-based。最大範囲を取得（空の場合は 0x0）
            max_row = ws.max_row or 0
            max_col = ws.max_column or 0
            # 0-based の2次元配列を構築（行・列とも 0 始まり）
            data: list[list[Any]] = []
            for r in range(1, max_row + 1):
                row_data: list[Any] = []
                for c in range(1, max_col + 1):
                    cell = ws.cell(row=r, column=c)
                    val = cell.value
                    # None は空文字にせずそのまま（フロントで未入力と区別しやすい）
                    row_data.append(val)
                data.append(row_data)

            # 結合セル情報を 0-based で取得（TreeView／簡易設計台の mergeCells 形式）
            merge_cells: list[dict[str, int]] = []
            merged_ranges = getattr(ws, "merged_cells", None)
            if merged_ranges is not None:
                ranges = getattr(merged_ranges, "ranges", [])
                for rng in ranges:
                    min_row = getattr(rng, "min_row", None)
                    max_row_r = getattr(rng, "max_row", None)
                    min_col = getattr(rng, "min_col", None)
                    max_col_r = getattr(rng, "max_col", None)
                    if None in (min_row, max_row_r, min_col, max_col_r):
                        continue
                    rowspan = max_row_r - min_row + 1
                    colspan = max_col_r - min_col + 1
                    if rowspan < 1 or colspan < 1:
                        continue
                    # 1-based → 0-based
                    merge_cells.append(
                        {
                            "row": min_row - 1,
                            "col": min_col - 1,
                            "rowspan": rowspan,
                            "colspan": colspan,
                        }
                    )

            # 行・列の非表示情報（0-based インデックスに対応）
            row_metadata: list[dict[str, bool]] = []
            for r in range(1, max_row + 1):
                rd = ws.row_dimensions.get(r)
                row_metadata.append(
                    {"hidden": getattr(rd, "hidden", False) if rd else False}
                )
            col_metadata: list[dict[str, Any]] = []
            for c in range(1, max_col + 1):
                letter = get_column_letter(c)
                cd = ws.column_dimensions.get(letter)
                raw_width = getattr(cd, "width", None) if cd else None
                width_chars = (
                    float(raw_width)
                    if raw_width is not None and raw_width > 0
                    else DEFAULT_COL_WIDTH_CHARS
                )
                width_px = int(round(width_chars * EXCEL_WIDTH_TO_PIXELS))
                col_metadata.append(
                    {
                        "hidden": getattr(cd, "hidden", False) if cd else False,
                        "width": width_px,
                    }
                )

            sheets_out.append(
                {
                    "name": sheet_name,
                    "data": data,
                    "mergeCells": merge_cells,
                    "row_metadata": row_metadata,
                    "col_metadata": col_metadata,
                }
            )
    finally:
        wb.close()

    return {"sheets": sheets_out}


# Excel の列幅: 未設定時のデフォルト（文字数）。OpenPyXL/Excel の目安。
DEFAULT_COL_WIDTH_CHARS = 8.43
# 列幅（文字数）→ ピクセル概算。目安: 1 文字 ≒ 7px（調整可能）
EXCEL_WIDTH_TO_PIXELS = 7


# OpenPyXL の有効範囲（1-based）
OPENPYXL_MIN_ROW = 1
OPENPYXL_MAX_ROW = 1048576
OPENPYXL_MIN_COL = 1
OPENPYXL_MAX_COL = 16384


def _normalize_cell_value(value: Any) -> str | int | float:
    """
    フロントから送られた value を OpenPyXL に安全に書き込める形に正規化する。
    None / 未設定 は空文字 "" に変換。数値はそのまま、それ以外は str に変換。
    """
    if value is None:
        return ""
    if isinstance(value, (str, int, float)):
        return value
    return str(value)


def _save_grid_via_xlwings(file_path: Path, changes: list[dict[str, Any]]) -> None:
    """
    xlwings で Excel ファイルを開き、changes のセル値だけを更新して保存する。
    画像・図形・グラフ等は触れずに保持される。Windows + Excel 前提。
    """
    import xlwings as xw

    path_str = str(file_path.resolve()).replace("/", "\\")
    app = xw.App(visible=False)
    try:
        book = app.books.open(path_str)
        try:
            for item in changes:
                sheet_name = item.get("sheetName")
                if not sheet_name:
                    continue
                try:
                    sheet = book.sheets[sheet_name]
                except (KeyError, Exception):
                    continue
                row1 = int(item.get("row", 0)) + 1
                col1 = int(item.get("col", 0)) + 1
                value = item.get("value")
                if value is None:
                    value = ""
                sheet.range((row1, col1)).value = value
            book.save()
        finally:
            book.close()
    finally:
        app.quit()


def _build_merged_covered_cells_per_sheet(wb: Any) -> dict[str, set[tuple[int, int]]]:
    """
    各シートについて、結合セル範囲の「左上以外」のセル (row1, col1) を 1-based で返す。
    書き込み禁止セル判定に使用する。
    """
    out: dict[str, set[tuple[int, int]]] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        covered: set[tuple[int, int]] = set()
        merged_ranges = getattr(ws, "merged_cells", None)
        if merged_ranges is not None:
            ranges = getattr(merged_ranges, "ranges", [])
            for rng in ranges:
                min_row = getattr(rng, "min_row", None)
                max_row_r = getattr(rng, "max_row", None)
                min_col = getattr(rng, "min_col", None)
                max_col_r = getattr(rng, "max_col", None)
                if None in (min_row, max_row_r, min_col, max_col_r):
                    continue
                for r in range(min_row, max_row_r + 1):
                    for c in range(min_col, max_col_r + 1):
                        if r != min_row or c != min_col:
                            covered.add((r, c))
        out[sheet_name] = covered
    return out


def save_grid(
    file_path: Path,
    changes: list[dict[str, Any]],
    use_excel_instance: bool = False,
) -> None:
    """
    指定されたセルだけ値を更新し、上書き保存する。
    スタイル（罫線・背景色・フォント等）は触れない。
    結合セル「左上以外」への書き込みは ValueError で拒否する。
    数式セルへの書き込み時は数式をクリアし、値を文字列として保存する。

    Args:
        file_path: テンプレートファイルのパス。
        changes: [ {"sheetName": str, "row": int, "col": int, "value": Any }, ... ]
                  row/col は 0-based（簡易設計台のグリッド形式）。value は文字列・数値・None など。
        use_excel_instance: True の場合、openpyxl では検証のみ行い、保存は xlwings（Excel 本体）
                           で行う。画像・図形が保持される。Windows + Excel 必須。
    """
    file_path = Path(file_path).resolve()
    if not file_path.exists():
        raise FileNotFoundError(f"テンプレートが見つかりません: {file_path}")

    if use_excel_instance and sys.platform != "win32":
        raise ValueError(
            "高忠実度保存（Excel本体を使用）は Windows 環境でのみ利用できます。"
        )

    wb = load_workbook(file_path, read_only=False, data_only=False)
    merged_covered = _build_merged_covered_cells_per_sheet(wb)

    # 共通バリデーション: 範囲チェック・結合セル書き込みチェック
    validated: list[dict[str, Any]] = []
    try:
        for item in changes:
            sheet_name = item.get("sheetName")
            if not sheet_name or sheet_name not in wb.sheetnames:
                continue

            try:
                row0 = int(item.get("row", 0))
                col0 = int(item.get("col", 0))
            except (TypeError, ValueError) as e:
                raise ValueError(
                    f"無効な座標です。row={item.get('row')}, col={item.get('col')}"
                ) from e

            row1 = row0 + 1
            col1 = col0 + 1

            if not (OPENPYXL_MIN_ROW <= row1 <= OPENPYXL_MAX_ROW):
                raise ValueError(
                    f"行番号が範囲外です。row (0-based)={row0} → (1-based)={row1}。"
                    f"有効範囲は {OPENPYXL_MIN_ROW} ～ {OPENPYXL_MAX_ROW} です。"
                )
            if not (OPENPYXL_MIN_COL <= col1 <= OPENPYXL_MAX_COL):
                raise ValueError(
                    f"列番号が範囲外です。col (0-based)={col0} → (1-based)={col1}。"
                    f"有効範囲は {OPENPYXL_MIN_COL} ～ {OPENPYXL_MAX_COL} です。"
                )

            covered = merged_covered.get(sheet_name, set())
            if (row1, col1) in covered:
                raise ValueError(
                    f"結合セルの一部のため書き込めません。"
                    f"シート={sheet_name}, 行={row0 + 1}, 列={col0 + 1}。左上のセルを指定してください。"
                )

            value = _normalize_cell_value(item.get("value"))
            validated.append(
                {"sheetName": sheet_name, "row": row0, "col": col0, "value": value}
            )

        if use_excel_instance:
            # openpyxl は保存せずに閉じる
            wb.close()
            wb = None
            # xlwings で開き、変更を適用して保存
            _save_grid_via_xlwings(file_path, validated)
        else:
            # 既存: openpyxl で変更を適用して保存
            for v in validated:
                sheet_name = v["sheetName"]
                row1 = v["row"] + 1
                col1 = v["col"] + 1
                value = v["value"]
                ws = wb[sheet_name]
                cell = ws.cell(row=row1, column=col1)
                was_formula = getattr(cell, "data_type", None) == "f"
                cell.value = value
                if was_formula:
                    cell.data_type = "s"
                elif isinstance(value, str):
                    cell.data_type = "s"
            wb.save(file_path)
    finally:
        if wb is not None:
            wb.close()
