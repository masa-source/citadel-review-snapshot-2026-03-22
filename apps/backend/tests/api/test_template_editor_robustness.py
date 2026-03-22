"""
設計台（Template Editor）堅牢化：構造的不適合の検知とテスト。

- 数式セルのプレースホルダ化
- 型変換の安全性（日付セルなど）
- 結合セル「左上以外」への書き込み拒否
- 非表示行・列メタデータの取得
"""

from __future__ import annotations

import sys
from datetime import date
from pathlib import Path

import pytest
import pytest_asyncio
from httpx import AsyncClient
from openpyxl import Workbook, load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

# バックエンドルートをパスに追加（import 前に必要）
_backend_root = Path(__file__).resolve().parents[2]
if str(_backend_root) not in sys.path:
    sys.path.insert(0, str(_backend_root))


def create_chaos_template_xlsx(dest_path: Path) -> None:
    """
    テスト用「カオス・テンプレート」を openpyxl で動的生成する。
    - 結合セル: A1:B2
    - 数式: C1 (=SUM(1,2))
    - 日付書式: D1 (2026/01/01)
    - 非表示行: 行5を非表示
    - シート保護: パスワードなしで有効化（特定セルのみロック想定）
    """
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("no active sheet")
    ws.title = "Sheet1"

    # 結合セル A1:B2
    ws.merge_cells("A1:B2")
    ws["A1"].value = "Merged"

    # 数式 C1
    ws["C1"].value = "=SUM(1,2)"

    # 日付 D1 (2026/01/01)
    ws["D1"].value = date(2026, 1, 1)
    ws["D1"].number_format = "YYYY/MM/DD"

    # 行5を非表示
    ws.row_dimensions[5].hidden = True
    ws["A5"].value = "HiddenRow"

    # シート保護（パスワードなしで有効化のみ。password は設定しない）
    ws.protection.sheet = True

    dest_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest_path)
    wb.close()


@pytest_asyncio.fixture
async def chaos_template_id(
    db_session: AsyncSession,
    client: AsyncClient,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    """
    カオス・テンプレートを tmp_path に作成し、DB に ReportTemplate を登録する。
    get_assets_base を tmp_path にパッチし、template_id を返す。
    """
    from tests.factories import insert_report_template

    # テンプレート配置: tmp_path / "templates" / "chaos_template.xlsx"
    template_dir = tmp_path / "templates"
    template_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = template_dir / "chaos_template.xlsx"
    create_chaos_template_xlsx(xlsx_path)

    # テンプレート・レポートルーターが参照する get_assets_base を tmp_path に差し替え
    monkeypatch.setattr("routers.templates.get_assets_base", lambda: tmp_path)
    monkeypatch.setattr("routers.reports.get_assets_base", lambda: tmp_path)

    # DB に ReportTemplate（テンプレート部品）を 1 件追加 (Factory経由)
    template = await insert_report_template(
        db_session, name="chaos_template", file_path="templates/chaos_template.xlsx"
    )
    template_id = template.id

    yield str(template_id)


# ---------------------------------------------------------------------------
# Test Case A: 数式のプレースホルダ化
# ---------------------------------------------------------------------------
@pytest.mark.asyncio
async def test_formula_cell_overwrite_with_placeholder(
    client: AsyncClient, chaos_template_id: str, tmp_path: Path
) -> None:
    """
    数式セル（C1）に {{ calculated_value }} を書き込む。
    保存後のファイルで C1 が数式ではなく文字列として保存されていることを検証する。
    """
    # フィクスチャでテンプレートを tmp_path に作成済み
    xlsx_path = tmp_path / "templates" / "chaos_template.xlsx"
    assert xlsx_path.exists()

    # C1 = 0-based で row=0, col=2
    response = await client.post(
        f"/api/templates/{chaos_template_id}/grid",
        json={
            "changes": [
                {
                    "sheetName": "Sheet1",
                    "row": 0,
                    "col": 2,
                    "value": "{{ calculated_value }}",
                }
            ],
        },
    )
    assert response.status_code == 200, response.text

    wb = load_workbook(xlsx_path, read_only=False, data_only=False)
    ws = wb.active
    assert ws is not None
    c1 = ws["C1"]
    assert c1.value == "{{ calculated_value }}", "C1 は文字列として保存されていること"
    assert getattr(c1, "data_type", None) == "s", (
        "C1 の data_type は文字列(s)であること"
    )
    wb.close()


# ---------------------------------------------------------------------------
# Test Case B: 型変換の安全性
# ---------------------------------------------------------------------------
@pytest.mark.asyncio
async def test_date_cell_overwrite_with_placeholder(
    client: AsyncClient, chaos_template_id: str, tmp_path: Path
) -> None:
    """
    日付セル（D1）に {{ date }} を書き込む。
    エラーにならず保存でき、セルのデータ型が適切に更新されていることを検証する。
    """
    xlsx_path = tmp_path / "templates" / "chaos_template.xlsx"
    assert xlsx_path.exists()

    # D1 = 0-based で row=0, col=3
    response = await client.post(
        f"/api/templates/{chaos_template_id}/grid",
        json={
            "changes": [
                {
                    "sheetName": "Sheet1",
                    "row": 0,
                    "col": 3,
                    "value": "{{ date }}",
                }
            ],
        },
    )
    assert response.status_code == 200, response.text

    wb = load_workbook(xlsx_path, read_only=False, data_only=False)
    ws = wb.active
    assert ws is not None
    d1 = ws["D1"]
    assert d1.value == "{{ date }}", "D1 は文字列として保存されていること"
    wb.close()


# ---------------------------------------------------------------------------
# Test Case C: 結合セルへの再書き込み（回帰テスト）
# ---------------------------------------------------------------------------
@pytest.mark.asyncio
async def test_merged_cell_part_rejected(
    client: AsyncClient, chaos_template_id: str
) -> None:
    """
    結合セル A1:B2 の「左上以外」（B2）への書き込みリクエストが拒否されることを検証する。
    """
    # B2 = 0-based で row=1, col=1
    response = await client.post(
        f"/api/templates/{chaos_template_id}/grid",
        json={
            "changes": [
                {
                    "sheetName": "Sheet1",
                    "row": 1,
                    "col": 1,
                    "value": "should be rejected",
                }
            ],
        },
    )
    assert response.status_code == 400, (
        "結合セル一部への書き込みは 400 で拒否されること"
    )
    data = response.json()
    assert "detail" in data
    detail = data["detail"] if isinstance(data["detail"], str) else str(data["detail"])
    assert "結合" in detail or "書き込めません" in detail or "左上" in detail


# ---------------------------------------------------------------------------
# Test Case D: 非表示領域の取得
# ---------------------------------------------------------------------------
@pytest.mark.asyncio
async def test_grid_includes_hidden_metadata(
    client: AsyncClient, chaos_template_id: str
) -> None:
    """
    GET /api/templates/{id}/grid のレスポンスに
    row_metadata / col_metadata が含まれ、非表示（hidden: true）の情報が
    正しく渡る構造になっていることを検証する（検知のみ）。
    """
    response = await client.get(f"/api/templates/{chaos_template_id}/grid")
    assert response.status_code == 200, response.text
    data = response.json()
    assert "sheets" in data
    sheets = data["sheets"]
    assert len(sheets) >= 1
    sheet = sheets[0]
    assert "row_metadata" in sheet, "row_metadata が含まれること"
    assert "col_metadata" in sheet, "col_metadata が含まれること"

    row_meta = sheet["row_metadata"]
    col_meta = sheet["col_metadata"]
    assert isinstance(row_meta, list), "row_metadata は配列であること"
    assert isinstance(col_meta, list), "col_metadata は配列であること"

    # カオステンプレートでは行5（0-based で index 4）が非表示
    assert len(row_meta) >= 5, "少なくとも行5までメタデータがあること"
    assert row_meta[4].get("hidden") is True, (
        "行5（0-based index 4）が hidden: true であること"
    )
