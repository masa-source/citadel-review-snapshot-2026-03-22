"""
POST /api/templates/auto-generate の API テスト。

テンプレート .xlsx をアップロードし、AI でレポート情報を抽出したうえで
プレースホルダを自動配置し、テンプレート部品として登録するフローを検証する。
- 通常 CI 用: extract_data_from_text をモックし、正常系で 200 と template/report を確認。
- 手動 AI 評価用: @pytest.mark.ai_eval で LM Studio 実機に投げて 200 を確認（CI では実行しない）。
"""

from io import BytesIO
from pathlib import Path
from unittest.mock import AsyncMock, patch

import pytest
import pytest_asyncio
from httpx import AsyncClient
from openpyxl import Workbook
from pydantic import ValidationError
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from ai_schemas import AIExtractedReport
from models import Report

# テストファイル基準の fixtures パス
FIXTURES_DIR = Path(__file__).resolve().parent.parent / "fixtures"
SAMPLE_COMPLEX_REPORT_XLSX = FIXTURES_DIR / "sample_complex_report.xlsx"


def _minimal_xlsx_bytes() -> bytes:
    """openpyxl で読み書きできる最小限の .xlsx を返す。"""
    wb = Workbook()
    ws = wb.active
    if ws is not None:
        ws["A1"] = "点検報告書"
        ws["A2"] = "テスト株式会社"
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


@pytest_asyncio.fixture
async def auto_generate_client(
    client: AsyncClient,
    db_session: AsyncSession,
    worker_tmp_dir: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    """auto-generate 用に assets パスを worker_tmp_dir に差し替えた client。"""
    monkeypatch.setattr(
        "routers.templates.get_assets_base",
        lambda: worker_tmp_dir,
    )
    monkeypatch.setattr(
        "routers.templates.get_assets_templates_dir",
        lambda: worker_tmp_dir / "templates",
    )
    (worker_tmp_dir / "templates").mkdir(parents=True, exist_ok=True)
    yield client


class TestTemplatesAutoGenerate:
    """POST /api/templates/auto-generate のテスト"""

    @pytest.mark.asyncio
    async def test_auto_generate_returns_200_with_template_and_report(
        self,
        auto_generate_client: AsyncClient,
    ) -> None:
        """正常系: 有効な .xlsx と name を送ると 200 で template と report が返る。"""
        xlsx_bytes = _minimal_xlsx_bytes()
        ai_data = AIExtractedReport(
            report_title="点検報告書",
            company_name="テスト株式会社",
        )

        with (
            patch(
                "routers.templates.extract_data_from_text",
                new_callable=AsyncMock,
                return_value=ai_data,
            ),
        ):
            response = await auto_generate_client.post(
                "/api/templates/auto-generate",
                files={
                    "file": (
                        "template.xlsx",
                        xlsx_bytes,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
                data={"name": "自動生成テンプレート"},
            )

        assert response.status_code == 200, response.text
        body = response.json()
        assert "template" in body
        assert "report" in body
        template = body["template"]
        report = body["report"]
        assert template["name"] == "自動生成テンプレート"
        assert "id" in template
        assert "filePath" in template
        assert report["reportTitle"] == "点検報告書"
        assert "id" in report

    @pytest.mark.asyncio
    async def test_auto_generate_without_file_returns_422(
        self,
        auto_generate_client: AsyncClient,
    ) -> None:
        """file なしで POST すると 422。"""
        response = await auto_generate_client.post(
            "/api/templates/auto-generate",
            data={"name": "OnlyName"},
        )
        assert response.status_code == 422

    @pytest.mark.asyncio
    async def test_auto_generate_when_ai_output_schema_invalid_returns_400(
        self,
        auto_generate_client: AsyncClient,
    ) -> None:
        xlsx_bytes = _minimal_xlsx_bytes()
        try:
            AIExtractedReport.model_validate(
                {"report_title": "x", "used_parts": [{"name": "A", "quantity": "two"}]}
            )
        except ValidationError as e:
            validation_error = e
        else:
            pytest.fail("expected ValidationError")

        with patch(
            "routers.templates.extract_data_from_text",
            new_callable=AsyncMock,
            side_effect=validation_error,
        ):
            response = await auto_generate_client.post(
                "/api/templates/auto-generate",
                files={
                    "file": (
                        "template.xlsx",
                        xlsx_bytes,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
                data={"name": "bad-ai"},
            )

        assert response.status_code == 400, response.text

    @pytest.mark.asyncio
    async def test_auto_generate_when_ai_raises_unexpected_error_returns_500(
        self,
        auto_generate_client: AsyncClient,
    ) -> None:
        xlsx_bytes = _minimal_xlsx_bytes()
        with patch(
            "routers.templates.extract_data_from_text",
            new_callable=AsyncMock,
            side_effect=RuntimeError("ai down"),
        ):
            response = await auto_generate_client.post(
                "/api/templates/auto-generate",
                files={
                    "file": (
                        "template.xlsx",
                        xlsx_bytes,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
                data={"name": "ai-down"},
            )
        assert response.status_code == 500, response.text

    @pytest.mark.asyncio
    async def test_auto_generate_when_late_failure_does_not_leave_report_row(
        self,
        auto_generate_client: AsyncClient,
        db_session: AsyncSession,
    ) -> None:
        xlsx_bytes = _minimal_xlsx_bytes()
        ai_data = AIExtractedReport(report_title="点検報告書")

        async def _build_report(session: AsyncSession, _ai_data: AIExtractedReport):
            r = Report(report_title=_ai_data.report_title)
            session.add(r)
            await session.flush()
            return r

        with (
            patch(
                "routers.templates.extract_data_from_text",
                new_callable=AsyncMock,
                return_value=ai_data,
            ),
            patch(
                "routers.templates.build_report_from_ai_data",
                new_callable=AsyncMock,
                side_effect=_build_report,
            ),
            patch(
                "routers.templates.shutil.copy2",
                side_effect=RuntimeError("disk full"),
            ),
        ):
            response = await auto_generate_client.post(
                "/api/templates/auto-generate",
                files={
                    "file": (
                        "late_fail.xlsx",
                        xlsx_bytes,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
                data={"name": "late-fail"},
            )
        assert response.status_code == 500, response.text

        row = await db_session.execute(
            select(Report).where(Report.report_title == "点検報告書")
        )
        assert row.scalars().all() == []

    @pytest.mark.asyncio
    async def test_auto_generate_with_sample_complex_report_xlsx_returns_200(
        self,
        auto_generate_client: AsyncClient,
    ) -> None:
        """正常系: fixtures/sample_complex_report.xlsx を送り、AI をモックして 200 で template/report が返る。"""
        if not SAMPLE_COMPLEX_REPORT_XLSX.exists():
            pytest.skip(f"fixture not found: {SAMPLE_COMPLEX_REPORT_XLSX}")
        xlsx_bytes = SAMPLE_COMPLEX_REPORT_XLSX.read_bytes()
        ai_data = AIExtractedReport(
            report_title="点検報告書",
            company_name="サンプル会社",
        )
        with (
            patch(
                "routers.templates.extract_data_from_text",
                new_callable=AsyncMock,
                return_value=ai_data,
            ),
        ):
            response = await auto_generate_client.post(
                "/api/templates/auto-generate",
                files={
                    "file": (
                        "sample_complex_report.xlsx",
                        xlsx_bytes,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
                data={"name": "サンプル複合レポートから生成"},
            )
        assert response.status_code == 200, response.text
        body = response.json()
        assert "template" in body
        assert "report" in body
        assert body["template"]["name"] == "サンプル複合レポートから生成"
        assert body["report"]["reportTitle"] == "点検報告書"

    @pytest.mark.ai_eval
    @pytest.mark.asyncio
    async def test_auto_generate_ai_eval_real_lm(
        self,
        auto_generate_client: AsyncClient,
    ) -> None:
        """手動 AI 評価: fixtures/sample_complex_report.xlsx を LM Studio 実機に投げて 200 を確認。CI では -m ai_eval を付けず実行するため、実機が使えない場合はスキップする。"""
        if not SAMPLE_COMPLEX_REPORT_XLSX.exists():
            pytest.skip(f"fixture not found: {SAMPLE_COMPLEX_REPORT_XLSX}")
        xlsx_bytes = SAMPLE_COMPLEX_REPORT_XLSX.read_bytes()
        response = await auto_generate_client.post(
            "/api/templates/auto-generate",
            files={
                "file": (
                    "sample_complex_report.xlsx",
                    xlsx_bytes,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            data={"name": "AI実機評価用テンプレート"},
        )
        if response.status_code >= 500:
            pytest.skip(
                f"LM Studio 実機が利用できないかエラー (status={response.status_code}): {response.text[:200]}"
            )
        assert response.status_code == 200, response.text
        body = response.json()
        assert "template" in body and "report" in body
        assert body["template"].get("name") == "AI実機評価用テンプレート"
        assert body["report"].get("reportTitle")
