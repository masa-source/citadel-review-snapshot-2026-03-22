from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook


def build_workbook() -> Workbook:
  """指定の3シート構成を持つテスト用 Excel ワークブックを生成する。"""

  wb = Workbook()

  # シート1: 表紙（基本情報）
  ws1 = wb.active
  ws1.title = "表紙"

  ws1["B2"] = "2026年度 ポンプ設備定期点検報告書"
  ws1["D2"] = "管理No. CTL-2026-001"

  ws1["B4"] = "顧客名: 株式会社シタデル工業"

  ws1["B5"] = "点検日: 2026/03/06"
  ws1["C5"] = "天候: 晴れ"

  ws1["B7"] = "作業責任者"
  ws1["C7"] = "山田 太郎"

  ws1["B8"] = "担当"
  ws1["C8"] = "佐藤花子"

  # シート2: 点検記録（対象計器）
  ws2 = wb.create_sheet("点検記録")

  ws2["B2"] = "(株)シタデル工業"

  ws2["B4"] = "Tag No."
  ws2["C4"] = "機器名称"
  ws2["D4"] = "型式"
  ws2["E4"] = "製造番号"
  ws2["F4"] = "判定"

  ws2["B5"] = "P-001"
  ws2["C5"] = "給水ポンプ"
  ws2["D5"] = "Model-A"
  ws2["E5"] = "SN-1111"
  ws2["F5"] = "良"

  ws2["B6"] = "P-002"
  ws2["C6"] = "循環ポンプ"
  ws2["D6"] = "Model-B"
  ws2["E6"] = "SN-2222"
  ws2["F6"] = "良"

  ws2["B7"] = "P-001"
  ws2["C7"] = "給水ポンプ(予備)"
  ws2["D7"] = "-"
  ws2["E7"] = "不明"
  ws2["F7"] = "停止"

  # B8 行は完全な空行（データを書き込まない）

  ws2["B9"] = "P-003"
  ws2["C9"] = "排水ポンプ"
  ws2["D9"] = "Model-C"
  ws2["E9"] = "SN-3333"
  ws2["F9"] = "良"

  # シート3: 交換部品リスト（使用部品）
  ws3 = wb.create_sheet("交換部品リスト")

  ws3["B2"] = "作業者確認: 山田太郎"

  ws3["B4"] = "部品名"
  ws3["C4"] = "型番"
  ws3["D4"] = "数量"
  ws3["E4"] = "備考"

  ws3["B5"] = "Oリング"
  ws3["C5"] = "OR-10"
  ws3["D5"] = 2
  ws3["E5"] = "ポンプA用"

  ws3["B6"] = "メカニカルシール"
  ws3["C6"] = "MS-50"
  ws3["D6"] = 1
  ws3["E6"] = ""

  # ノイズデータ（AIには無視してほしい行）
  ws3["B7"] = "なし"
  ws3["C7"] = "-"
  ws3["D7"] = 0
  ws3["E7"] = ""

  return wb


def main() -> None:
  project_root = Path(__file__).resolve().parent.parent
  output_path = (
    project_root / "apps" / "backend" / "tests" / "fixtures" / "sample_complex_report.xlsx"
  )
  output_path.parent.mkdir(parents=True, exist_ok=True)

  wb = build_workbook()
  wb.save(output_path)
  print(f"Generated: {output_path}")


if __name__ == "__main__":
  main()

